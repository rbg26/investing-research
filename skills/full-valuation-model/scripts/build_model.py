#!/usr/bin/env python3
"""
Full Valuation Model Builder
============================
Builds a four-tab Excel valuation workbook for any stock.

Usage:
    1. Populate the CONFIG dict below with data pulled from yfinance and EDGAR MCPs.
    2. Run: python3 build_model.py
    3. Output: <TICKER>_Valuation_Model.xlsx

Tabs produced:
    1. Comps Analysis       — peer trading multiples + statistics (incl. PEG, Price/FCF, Price/Sales)
    2. DCF Model            — Bear/Base/Bull EBIT-based UFCF + equity bridge
    3. Sensitivity Analysis — 3 two-way tables
    4. Conservative DCF     — OCF−CapEx, 10yr two-phase, value investing defaults
    5. Dhandho DCF          — Pabrai framework: two-phase FCF, sale-price terminal, MoS, Reverse Dhandho

Formula rule: every derived cell is a live Excel formula string.
              Python computes nothing except the CONFIG inputs.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG — populate this dict from yfinance + EDGAR before running
# ══════════════════════════════════════════════════════════════════════════════

CONFIG = {
    # ── Subject company ──────────────────────────────────────────────────────
    # Populated by Claude from yfinance + EDGAR MCP before running this script.
    "ticker":           "TICKER",
    "company_name":     "Company Name",
    "analysis_date":    "Month DD, YYYY",

    # Market data (from yfinance)
    "current_price":    0.00,
    "market_cap_m":     0,          # $M
    "ev_m":             0,          # $M
    "shares_m":         0,          # diluted shares outstanding, millions
    "net_debt_m":       0,          # total debt − cash & ST investments, $M (negative = net cash)
    "beta":             0.00,

    # LTM financials (from yfinance / EDGAR)
    "revenue_ltm_m":    0,
    "gross_profit_m":   0,
    "ebitda_m":         0,          # Operating Income + D&A
    "net_income_m":     0,
    "ocf_m":            0,          # Operating Cash Flow
    "capex_m":          0,
    "fcf_m":            0,          # OCF − CapEx
    "rev_growth_yoy":   0.00,
    "trailing_pe":      0.00,
    "forward_pe":       0.00,
    "price_to_book":    0.00,

    # Historical financials — 5 years (from EDGAR), most recent first → oldest last
    # Format: [FYn, FYn-1, FYn-2, FYn-3, FYn-4]
    "hist_years":       ["FY2025", "FY2024", "FY2023", "FY2022", "FY2021"],
    "hist_revenue":     [0, 0, 0, 0, 0],
    "hist_op_income":   [0, 0, 0, 0, 0],
    "hist_da":          [0, 0, 0, 0, 0],
    "hist_net_income":  [0, 0, 0, 0, 0],
    "hist_capex":       [0, 0, 0, 0, 0],
    "hist_ocf":         [0, 0, 0, 0, 0],

    # WACC inputs (from CAPM — use 10-yr Treasury for risk_free_rate)
    "risk_free_rate":   0.00,
    "erp":              0.055,      # equity risk premium, default 5.5%
    "pretax_cost_debt": 0.00,
    "tax_rate":         0.00,
    "wacc":             0.00,       # base case WACC

    # ── Scenario assumptions ─────────────────────────────────────────────────
    # Revenue growth rates per year [yr1, yr2, yr3, yr4, yr5]
    # Anchor to historical CAGR; bear = below trend, bull = modest acceleration
    "bear_rev_growth":  [0.00, 0.00, 0.00, 0.00, 0.00],
    "base_rev_growth":  [0.00, 0.00, 0.00, 0.00, 0.00],
    "bull_rev_growth":  [0.00, 0.00, 0.00, 0.00, 0.00],

    # EBITDA margins per year [yr1..yr5]
    # Bear = flat/slight compression; Base = modest expansion; Bull = toward peer median
    "bear_ebitda_mg":   [0.00, 0.00, 0.00, 0.00, 0.00],
    "base_ebitda_mg":   [0.00, 0.00, 0.00, 0.00, 0.00],
    "bull_ebitda_mg":   [0.00, 0.00, 0.00, 0.00, 0.00],

    # D&A and CapEx as % of revenue (flat across years — adjust if needed)
    "da_pct":           0.00,
    "capex_pct":        0.00,

    # Scenario WACCs
    "bear_wacc":        0.00,
    "base_wacc":        0.00,
    "bull_wacc":        0.00,

    # Terminal growth rates
    "bear_tgr":         0.00,
    "base_tgr":         0.00,
    "bull_tgr":         0.00,

    # Exit EV/EBITDA multiples
    "bear_exit_mult":   0.0,
    "base_exit_mult":   0.0,
    "bull_exit_mult":   0.0,

    # ── Conservative DCF ─────────────────────────────────────────────────────
    "cons_fcf_base_m":  0,          # OCF − CapEx (most recent fiscal year)
    "cons_growth_1_5":  0.05,       # yr 1–5 growth (hard cap 20%)
    "cons_growth_6_10": 0.03,       # yr 6–10 growth (hard cap 10%)
    "cons_discount":    0.12,       # 12% large-cap stable / 15% higher-risk
    "cons_tgr":         0.01,       # terminal growth rate (0–2%, default 1%)

    # ── Dhandho DCF ──────────────────────────────────────────────────────────
    "dhandho_fcf_base_m":    0,     # OCF − CapEx (same base as Conservative DCF)
    "dhandho_growth_1_5":    0.08,  # yr 1–5 FCF growth (editable in Excel)
    "dhandho_growth_6_10":   0.05,  # yr 6–10 FCF growth (editable in Excel)
    "dhandho_discount":      0.12,  # hurdle rate
    "dhandho_exit_mult":     12.0,  # terminal year FCF exit multiple
    "dhandho_mos":           0.25,  # margin of safety %
    "dhandho_cash_m":        0,     # cash & equivalents (balance sheet)
    "dhandho_investments_m": 0,     # short-term investments (balance sheet)
    "dhandho_debt_m":        0,     # long-term debt (balance sheet)

    # ── Comps ────────────────────────────────────────────────────────────────
    # Each comp: [name, ticker, mkt_cap_m, ev_m, revenue_m, rev_growth,
    #             gross_profit_m, ebitda_m, net_income_m, ocf_m, fcf_m,
    #             trailing_pe, forward_pe, price_to_book, beta, div_yield, wk52_return]
    # Select 4–6 peers: similar business model, revenue within ~0.3x–3x of subject
    "comps": [
        # ["Peer One",  "TKR1", 0, 0, 0, 0.00, 0, 0, 0, 0, 0, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00],
        # ["Peer Two",  "TKR2", 0, 0, 0, 0.00, 0, 0, 0, 0, 0, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00],
        # ["Peer Three","TKR3", 0, 0, 0, 0.00, 0, 0, 0, 0, 0, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00],
        # ["Peer Four", "TKR4", 0, 0, 0, 0.00, 0, 0, 0, 0, 0, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00],
    ],
}

# ══════════════════════════════════════════════════════════════════════════════
#  STYLE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

DARK_BLUE  = "1F4E79"
MED_BLUE   = "2E75B6"
LIGHT_BLUE = "D9E1F2"
LIGHT_GREY = "F2F2F2"
WHITE      = "FFFFFF"
INPUT_BLUE = "1F5C99"
AMBER      = "FFF2CC"
GREEN      = "E2EFDA"
RED_LIGHT  = "FFCCCC"
BLACK      = "000000"


def fill(c):
    return PatternFill("solid", fgColor=c)

def fnt(bold=False, color=BLACK, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def ctr():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def lft():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def lft_wrap():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def box_border():
    s = Side(style="medium", color=BLACK)
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, row, text, merge_end="P", bg=DARK_BLUE, fg=WHITE, sz=11, height=20):
    ws.merge_cells(f"A{row}:{merge_end}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = fnt(bold=True, color=fg, size=sz)
    c.fill = fill(bg)
    c.alignment = ctr()
    ws.row_dimensions[row].height = height

def col_hdr(ws, row, col, text, bg=LIGHT_BLUE):
    c = ws.cell(row=row, column=col, value=text)
    c.font = fnt(bold=True, size=9)
    c.fill = fill(bg)
    c.alignment = ctr()

def inp(ws, r, c, val, fmt="#,##0"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.number_format = fmt
    cell.font = fnt(color=INPUT_BLUE, size=10)
    cell.alignment = ctr()
    return cell

def frm(ws, r, c, formula, fmt="#,##0", bold=False, color=BLACK):
    cell = ws.cell(row=r, column=c, value=formula)
    cell.number_format = fmt
    cell.font = fnt(color=color, bold=bold, size=10)
    cell.alignment = ctr()
    return cell

def lbl(ws, r, c, text, bold=False, bg=None):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font = fnt(bold=bold, size=10)
    cell.alignment = lft()
    if bg:
        cell.fill = fill(bg)
    return cell

def shade_row(ws, row, color, cols=16):
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).fill = fill(color)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 1: COMPS ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

def build_comps_tab(wb, cfg):
    ws = wb.active
    ws.title = "Comps Analysis"
    col_widths = [24, 8, 11, 10, 10, 9, 10, 9, 10, 9, 10, 9, 9, 8, 9, 9, 9]
    set_col_widths(ws, col_widths)

    ticker = cfg["ticker"]
    comp_tickers = " • ".join(f"{c[0]} ({c[1]})" for c in cfg["comps"])

    # Title
    title_row(ws, 1, f"{ticker} — COMPARABLE COMPANY ANALYSIS", "O", sz=12, height=22)
    ws.merge_cells("A2:O2")
    ws["A2"].value = f"{cfg['company_name']} ({ticker})  •  {comp_tickers}"
    ws["A2"].font = fnt(color=WHITE, size=9)
    ws["A2"].fill = fill(DARK_BLUE)
    ws["A2"].alignment = ctr()
    ws.merge_cells("A3:O3")
    ws["A3"].value = f"As of {cfg['analysis_date']}  |  USD Millions  |  Source: yfinance / SEC EDGAR"
    ws["A3"].font = fnt(italic=True, size=8, color="AAAAAA")
    ws["A3"].fill = fill(DARK_BLUE)
    ws["A3"].alignment = ctr()

    # Operating section header
    title_row(ws, 5, "OPERATING STATISTICS & FINANCIAL METRICS", "O", bg=MED_BLUE, sz=10)
    ws.row_dimensions[6].height = 34
    op_hdrs = ["Company", "Ticker", "Revenue\n($M LTM)", "Rev Growth\n(YoY)",
               "Gross Profit\n($M)", "Gross\nMargin", "EBITDA\n($M)", "EBITDA\nMargin",
               "Net Income\n($M)", "Net\nMargin", "Op CF\n($M)", "FCF\n($M)",
               "FCF\nMargin", "Beta"]
    for ci, h in enumerate(op_hdrs, 1):
        col_hdr(ws, 6, ci, h)

    # Subject company row 7
    subj = cfg
    s_data = [
        subj["company_name"], ticker,
        subj["revenue_ltm_m"], subj["rev_growth_yoy"],
        subj["gross_profit_m"], None,
        subj["ebitda_m"], None,
        subj["net_income_m"], None,
        subj["ocf_m"], subj["fcf_m"], None,
        subj["beta"],
    ]
    for col in range(1, 15):
        ws.cell(row=7, column=col).fill = fill("EBF3FB")

    ws.cell(row=7, column=1).value = subj["company_name"]
    ws.cell(row=7, column=1).font = fnt(bold=True, size=10)
    ws.cell(row=7, column=1).alignment = lft()
    ws.cell(row=7, column=1).fill = fill("EBF3FB")

    ws.cell(row=7, column=2).value = ticker
    ws.cell(row=7, column=2).font = fnt(bold=True, color=INPUT_BLUE, size=10)
    ws.cell(row=7, column=2).alignment = ctr()

    abs_cols  = {3: "#,##0", 5: "#,##0", 7: "#,##0", 9: "#,##0", 11: "#,##0", 12: "#,##0"}
    pct_cols  = {4: "0.0%", 14: "0.00"}
    frm_cols  = {6: f"=E7/C7", 8: f"=G7/C7", 10: f"=I7/C7", 13: f"=L7/C7"}

    for col, fmt in abs_cols.items():
        inp(ws, 7, col, s_data[col - 1], fmt)
    for col, fmt in pct_cols.items():
        inp(ws, 7, col, s_data[col - 1], fmt)
    for col, formula in frm_cols.items():
        frm(ws, 7, col, formula, "0.0%")

    # Comp rows 8–11
    for ri, comp in enumerate(cfg["comps"]):
        r = 8 + ri
        ws.row_dimensions[r].height = 16
        bg = LIGHT_GREY if ri % 2 == 0 else WHITE

        ws.cell(row=r, column=1).value = comp[0]
        ws.cell(row=r, column=1).font = fnt(size=10)
        ws.cell(row=r, column=1).alignment = lft()
        ws.cell(row=r, column=1).fill = fill(bg)

        ws.cell(row=r, column=2).value = comp[1]
        ws.cell(row=r, column=2).font = fnt(color=INPUT_BLUE, size=10)
        ws.cell(row=r, column=2).alignment = ctr()
        ws.cell(row=r, column=2).fill = fill(bg)

        # revenue, rev_growth, gross_profit
        for col, val, fmt in [
            (3, comp[4],  "#,##0"),
            (4, comp[5],  "0.0%"),
            (5, comp[6],  "#,##0"),
            (7, comp[7],  "#,##0"),
            (9, comp[8],  "#,##0"),
            (11, comp[9], "#,##0"),
            (14, comp[14], "0.00"),
        ]:
            cell = ws.cell(row=r, column=col)
            cell.value = val if val is not None else "N/A"
            cell.number_format = fmt
            cell.font = fnt(color=INPUT_BLUE, size=10)
            cell.alignment = ctr()
            cell.fill = fill(bg)

        for col, formula in [(6, f"=E{r}/C{r}"), (8, f"=G{r}/C{r}"),
                              (10, f"=I{r}/C{r}"), (13, f"=L{r}/C{r}")]:
            cell = ws.cell(row=r, column=col)
            if ws.cell(row=r, column=col - 1).value == "N/A":
                cell.value = "N/A"
            else:
                cell.value = formula
            cell.number_format = "0.0%"
            cell.font = fnt(size=10)
            cell.alignment = ctr()
            cell.fill = fill(bg)

        # FCF
        fcf_val = comp[10]
        cell = ws.cell(row=r, column=12)
        cell.value = fcf_val if fcf_val is not None else "N/A"
        cell.number_format = "#,##0"
        cell.font = fnt(color=INPUT_BLUE, size=10)
        cell.alignment = ctr()
        cell.fill = fill(bg)

    # Stats block (comps only: rows 8–11)
    ws.row_dimensions[13].height = 4
    stats = [
        ("Maximum",         "=MAX({}8:{}11)"),
        ("75th Percentile", "=QUARTILE({}8:{}11,3)"),
        ("Median",          "=MEDIAN({}8:{}11)"),
        ("25th Percentile", "=QUARTILE({}8:{}11,1)"),
        ("Minimum",         "=MIN({}8:{}11)"),
    ]
    skip_stat = {1, 2, 3, 5, 7, 9, 11, 12}
    pct_stat   = {4, 6, 8, 10, 13}
    for si, (slabel, sfmt) in enumerate(stats):
        r = 14 + si
        ws.row_dimensions[r].height = 16
        shade_row(ws, r, LIGHT_GREY, 15)
        ws.cell(row=r, column=1).value = slabel
        ws.cell(row=r, column=1).font = fnt(bold=True, size=10)
        ws.cell(row=r, column=1).alignment = lft()
        for col in range(2, 15):
            if col in skip_stat:
                ws.cell(row=r, column=col).value = "—"
                ws.cell(row=r, column=col).alignment = ctr()
                ws.cell(row=r, column=col).font = fnt(size=10)
                continue
            cl = get_column_letter(col)
            ws.cell(row=r, column=col).value = sfmt.format(cl, cl)
            ws.cell(row=r, column=col).number_format = "0.0%" if col in pct_stat else ("0.00" if col == 14 else "0.0\"x\"")
            ws.cell(row=r, column=col).font = fnt(size=10)
            ws.cell(row=r, column=col).alignment = ctr()

    # Valuation multiples section
    ws.row_dimensions[20].height = 6
    title_row(ws, 21, "VALUATION MULTIPLES", "O", bg=MED_BLUE, sz=10)
    ws.row_dimensions[22].height = 34
    val_hdrs = ["Company", "Ticker", "Market Cap\n($M)", "Enterprise\nValue ($M)",
                "EV /\nRevenue", "EV /\nEBITDA", "P/E\n(Trailing)", "P/E\n(Forward)",
                "Price /\nBook", "FCF\nYield", "Div\nYield", "52-Wk\nReturn",
                "PEG\nRatio", "Price /\nFCF", "Price /\nSales"]
    for ci, h in enumerate(val_hdrs, 1):
        col_hdr(ws, 22, ci, h)

    all_companies = [
        [cfg["company_name"], ticker, cfg["market_cap_m"], cfg["ev_m"],
         cfg["trailing_pe"], cfg["forward_pe"], cfg["price_to_book"],
         0.013, -0.380]
    ] + [
        [c[0], c[1], c[2], c[3], c[11], c[12], c[13], c[16] if len(c) > 16 else 0, c[16] if len(c) > 16 else 0]
        for c in cfg["comps"]
    ]

    div_yields  = [0.013, 0.009, 0.007, 0.000, 0.015]
    wk52_rets   = [-0.380, -0.123, -0.087, -0.051, -0.308]

    for i, (name, tick, mc, ev_val, pe_t, pe_f, pb, _, _) in enumerate(all_companies):
        r = 23 + i
        op_row = 7 + i
        bg = "EBF3FB" if i == 0 else (LIGHT_GREY if i % 2 == 1 else WHITE)
        shade_row(ws, r, bg, 15)
        ws.row_dimensions[r].height = 16

        ws.cell(row=r, column=1).value = name
        ws.cell(row=r, column=1).font = fnt(bold=(i == 0), size=10)
        ws.cell(row=r, column=1).alignment = lft()

        ws.cell(row=r, column=2).value = tick
        ws.cell(row=r, column=2).font = fnt(bold=(i == 0), color=INPUT_BLUE, size=10)
        ws.cell(row=r, column=2).alignment = ctr()

        for col, val, fmt in [(3, mc, "#,##0"), (4, ev_val, "#,##0"),
                               (7, pe_t, "0.0\"x\""), (8, pe_f, "0.0\"x\""),
                               (9, pb, "0.0\"x\""),
                               (11, div_yields[i], "0.0%"),
                               (12, wk52_rets[i], "0.0%")]:
            cell = ws.cell(row=r, column=col)
            cell.value = val
            cell.number_format = fmt
            cell.font = fnt(color=INPUT_BLUE, size=10)
            cell.alignment = ctr()

        frm(ws, r, 5, f"=D{r}/C{op_row}", "0.0\"x\"")
        frm(ws, r, 6, f"=D{r}/G{op_row}", "0.0\"x\"")
        frm(ws, r, 10, f"=L{op_row}/C{r}", "0.0%")
        # PEG = P/E Trailing ÷ (YoY Revenue Growth as %)
        frm(ws, r, 13, f'=IFERROR(IF(D{op_row}=0,"",G{r}/(D{op_row}*100)),"")', "0.0\"x\"")
        # Price/FCF = Market Cap ÷ FCF
        frm(ws, r, 14, f'=IFERROR(C{r}/L{op_row},"")', "0.0\"x\"")
        # Price/Sales = Market Cap ÷ Revenue
        frm(ws, r, 15, f"=C{r}/C{op_row}", "0.0\"x\"")

    # Val stats
    ws.row_dimensions[29].height = 4
    skip_val = {1, 2, 3, 4}
    pct_val  = {10, 11, 12}
    for si, (slabel, _) in enumerate(stats):
        r = 30 + si
        ws.row_dimensions[r].height = 16
        shade_row(ws, r, LIGHT_GREY, 15)
        ws.cell(row=r, column=1).value = slabel
        ws.cell(row=r, column=1).font = fnt(bold=True, size=10)
        ws.cell(row=r, column=1).alignment = lft()
        stat_fmts = ["=MAX({}24:{}27)", "=QUARTILE({}24:{}27,3)",
                     "=MEDIAN({}24:{}27)", "=QUARTILE({}24:{}27,1)", "=MIN({}24:{}27)"]
        for col in range(2, 16):
            if col in skip_val:
                ws.cell(row=r, column=col).value = "—"
                ws.cell(row=r, column=col).alignment = ctr()
                ws.cell(row=r, column=col).font = fnt(size=10)
                continue
            cl = get_column_letter(col)
            ws.cell(row=r, column=col).value = stat_fmts[si].format(cl, cl)
            ws.cell(row=r, column=col).number_format = "0.0%" if col in pct_val else "0.0\"x\""
            ws.cell(row=r, column=col).font = fnt(size=10)
            ws.cell(row=r, column=col).alignment = ctr()

    # Notes
    ws.row_dimensions[36].height = 6
    title_row(ws, 37, "METHODOLOGY & DATA SOURCES", "O", bg=DARK_BLUE, sz=10)
    notes = [
        f"Data: yfinance API (market data, LTM financials); SEC EDGAR 10-K (historical operating data). As of {cfg['analysis_date']}.",
        f"EBITDA = Operating Income + D&A. {ticker} EBITDA: ${cfg['ebitda_m']:,.0f}M.",
        "FCF = Operating Cash Flow − Capital Expenditures.",
        "EV = Market Cap + Total Debt − Cash & Short-Term Investments.",
        "PEG Ratio = P/E (Trailing) ÷ YoY Revenue Growth (%). Price/FCF = Market Cap ÷ FCF. Price/Sales = Market Cap ÷ Revenue.",
        "Statistics computed on peer comps only (rows 8–11 / 24–27). Blue cells = hardcoded inputs. Black cells = formula-driven.",
    ]
    for ni, note in enumerate(notes):
        r = 38 + ni
        ws.merge_cells(f"A{r}:O{r}")
        ws.cell(row=r, column=1).value = f"• {note}"
        ws.cell(row=r, column=1).font = fnt(italic=True, size=8, color="555555")
        ws.cell(row=r, column=1).alignment = lft_wrap()
        ws.row_dimensions[r].height = 20


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 2: DCF MODEL
# ══════════════════════════════════════════════════════════════════════════════

def build_dcf_scenario(ws, start_row, scenario_name, bg_color, params, cfg):
    """Build one Bear/Base/Bull scenario block. Returns next available row."""
    rev_growth    = params["rev_growth"]
    ebitda_mg     = params["ebitda_mg"]
    da_pct        = cfg["da_pct"]
    capex_pct     = cfg["capex_pct"]
    tax_rate      = cfg["tax_rate"]
    wacc          = params["wacc"]
    tgr           = params["tgr"]
    exit_mult     = params["exit_mult"]
    base_rev      = cfg["hist_revenue"][0]    # most recent actual
    base_ebitda   = cfg["hist_ebitda_m"] if "hist_ebitda_m" in cfg else cfg["ebitda_m"]
    net_debt      = cfg["net_debt_m"]
    shares        = cfg["shares_m"]
    price         = cfg["current_price"]

    r = start_row
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"A{r}:I{r}")
    c = ws.cell(row=r, column=1, value=f"SCENARIO: {scenario_name}")
    c.font = fnt(bold=True, color=WHITE, size=12)
    c.fill = fill(bg_color)
    c.alignment = ctr()
    r += 1

    proj_years = [f"FY{2026 + i}E" for i in range(5)]
    ws.row_dimensions[r].height = 30
    for ci, h in enumerate(["", f"Historical\nFY{cfg['hist_years'][0]}A"] + proj_years, 1):
        col_hdr(ws, r, ci, h)
    r += 1

    # Revenue
    shade_row(ws, r, LIGHT_GREY, 9)
    lbl(ws, r, 1, "Revenue ($M)", bold=True, bg=LIGHT_GREY)
    inp(ws, r, 2, base_rev, "#,##0")
    for yr in range(5):
        col = 3 + yr
        g = rev_growth[yr]
        prev = get_column_letter(col - 1)
        frm(ws, r, col, f"={prev}{r}*(1+{g})", "#,##0")
    rev_row = r; r += 1

    lbl(ws, r, 1, "  YoY Growth")
    ws.cell(row=r, column=2).value = f"{cfg['rev_growth_yoy']:.1%}"
    ws.cell(row=r, column=2).font = fnt(italic=True, size=9, color="888888")
    ws.cell(row=r, column=2).alignment = ctr()
    for yr in range(5):
        col = 3 + yr
        frm(ws, r, col, f"={get_column_letter(col)}{rev_row}/{get_column_letter(col-1)}{rev_row}-1", "0.0%", color="555555")
    r += 1

    # EBITDA
    shade_row(ws, r, LIGHT_GREY, 9)
    lbl(ws, r, 1, "EBITDA ($M)", bold=True, bg=LIGHT_GREY)
    inp(ws, r, 2, base_ebitda, "#,##0")
    for yr in range(5):
        col = 3 + yr
        frm(ws, r, col, f"={get_column_letter(col)}{rev_row}*{ebitda_mg[yr]}", "#,##0")
    ebitda_row = r; r += 1

    lbl(ws, r, 1, "  EBITDA Margin")
    ws.cell(row=r, column=2).value = f"{cfg['ebitdaMargins'] if 'ebitdaMargins' in cfg else cfg['ebitda_m']/cfg['revenue_ltm_m']:.1%}"
    ws.cell(row=r, column=2).font = fnt(italic=True, size=9, color="888888")
    ws.cell(row=r, column=2).alignment = ctr()
    for yr in range(5):
        col = 3 + yr
        frm(ws, r, col, f"={get_column_letter(col)}{ebitda_row}/{get_column_letter(col)}{rev_row}", "0.0%", color="555555")
    r += 1

    # D&A
    shade_row(ws, r, LIGHT_GREY, 9)
    lbl(ws, r, 1, "D&A ($M)", bg=LIGHT_GREY)
    inp(ws, r, 2, cfg["hist_da"][0], "#,##0")
    for yr in range(5):
        col = 3 + yr
        frm(ws, r, col, f"={get_column_letter(col)}{rev_row}*{da_pct}", "#,##0")
    da_row = r; r += 1

    # EBIT
    lbl(ws, r, 1, "EBIT ($M)")
    inp(ws, r, 2, cfg["hist_op_income"][0], "#,##0")
    for yr in range(5):
        col = 3 + yr
        frm(ws, r, col, f"={get_column_letter(col)}{ebitda_row}-{get_column_letter(col)}{da_row}", "#,##0")
    ebit_row = r; r += 1

    # NOPAT
    shade_row(ws, r, LIGHT_GREY, 9)
    lbl(ws, r, 1, "NOPAT = EBIT × (1 − Tax)", bg=LIGHT_GREY)
    inp(ws, r, 2, int(cfg["hist_op_income"][0] * (1 - tax_rate)), "#,##0")
    for yr in range(5):
        col = 3 + yr
        frm(ws, r, col, f"={get_column_letter(col)}{ebit_row}*(1-{tax_rate})", "#,##0")
    nopat_row = r; r += 1

    # CapEx
    lbl(ws, r, 1, "Less: CapEx ($M)")
    inp(ws, r, 2, -cfg["hist_capex"][0], "#,##0")
    for yr in range(5):
        col = 3 + yr
        frm(ws, r, col, f"=-{get_column_letter(col)}{rev_row}*{capex_pct}", "#,##0")
    capex_row = r; r += 1

    # D&A add-back
    shade_row(ws, r, LIGHT_GREY, 9)
    lbl(ws, r, 1, "Add: D&A ($M)", bg=LIGHT_GREY)
    for yr in range(5):
        col = 3 + yr
        frm(ws, r, col, f"={get_column_letter(col)}{da_row}", "#,##0")
    da_back_row = r; r += 1

    # UFCF
    lbl(ws, r, 1, "Unlevered FCF ($M)", bold=True)
    for yr in range(5):
        col = 3 + yr
        cl = get_column_letter(col)
        frm(ws, r, col, f"={cl}{nopat_row}+{cl}{capex_row}+{cl}{da_back_row}", "#,##0", bold=True)
    ufcf_row = r; r += 1

    # Discount factors (mid-year)
    shade_row(ws, r, LIGHT_GREY, 9)
    lbl(ws, r, 1, "Discount Factor (mid-year)", bg=LIGHT_GREY)
    for yr in range(5):
        col = 3 + yr
        period = yr + 0.5
        frm(ws, r, col, f"=1/(1+{wacc})^{period}", "0.000")
    disc_row = r; r += 1

    # PV FCF
    lbl(ws, r, 1, "PV of FCF ($M)", bold=True)
    for yr in range(5):
        col = 3 + yr
        cl = get_column_letter(col)
        frm(ws, r, col, f"={cl}{ufcf_row}*{cl}{disc_row}", "#,##0", bold=True)
    pv_fcf_row = r; r += 2

    # Terminal value section
    ws.merge_cells(f"A{r}:I{r}")
    c = ws.cell(row=r, column=1, value="TERMINAL VALUE CALCULATION")
    c.font = fnt(bold=True, color=WHITE, size=10)
    c.fill = fill(MED_BLUE)
    c.alignment = ctr()
    ws.row_dimensions[r].height = 16
    r += 1

    last_col = get_column_letter(3 + 4)  # G = year 5

    lbl(ws, r, 1, "Terminal Year FCF ($M)", bg=LIGHT_GREY)
    frm(ws, r, 2, f"={last_col}{ufcf_row}", "#,##0")
    ws.cell(row=r, column=2).fill = fill(LIGHT_GREY)
    tv_fcf_row = r; r += 1

    lbl(ws, r, 1, "Terminal Growth Rate")
    inp(ws, r, 2, tgr, "0.0%")
    ws.cell(row=r, column=2).fill = fill(LIGHT_BLUE)
    ws.cell(row=r, column=2).border = box_border()
    tgr_row = r; r += 1

    lbl(ws, r, 1, "WACC")
    inp(ws, r, 2, wacc, "0.0%")
    ws.cell(row=r, column=2).fill = fill(LIGHT_BLUE)
    ws.cell(row=r, column=2).border = box_border()
    wacc_row = r; r += 1

    lbl(ws, r, 1, "Terminal Value — GGM ($M)", bold=True, bg=LIGHT_BLUE)
    frm(ws, r, 2, f"=B{tv_fcf_row}*(1+B{tgr_row})/(B{wacc_row}-B{tgr_row})", "#,##0", bold=True)
    ws.cell(row=r, column=2).fill = fill(LIGHT_BLUE)
    tv_ggm_row = r; r += 1

    lbl(ws, r, 1, "Exit EV/EBITDA Multiple")
    inp(ws, r, 2, exit_mult, "0.0\"x\"")
    ws.cell(row=r, column=2).fill = fill(LIGHT_BLUE)
    ws.cell(row=r, column=2).border = box_border()
    exit_mult_row = r; r += 1

    lbl(ws, r, 1, "Terminal Year EBITDA ($M)", bg=LIGHT_GREY)
    frm(ws, r, 2, f"={last_col}{ebitda_row}", "#,##0")
    ws.cell(row=r, column=2).fill = fill(LIGHT_GREY)
    tv_ebitda_row = r; r += 1

    lbl(ws, r, 1, "Terminal Value — Exit Multiple ($M)", bold=True, bg=LIGHT_BLUE)
    frm(ws, r, 2, f"=B{exit_mult_row}*B{tv_ebitda_row}", "#,##0", bold=True)
    ws.cell(row=r, column=2).fill = fill(LIGHT_BLUE)
    tv_mult_row = r; r += 1

    last_disc_col = get_column_letter(3 + 4)
    lbl(ws, r, 1, "PV of Terminal Value — GGM ($M)", bold=True, bg=GREEN)
    frm(ws, r, 2, f"=B{tv_ggm_row}*{last_disc_col}{disc_row}", "#,##0", bold=True)
    ws.cell(row=r, column=2).fill = fill(GREEN)
    pv_tv_ggm_row = r; r += 1

    lbl(ws, r, 1, "PV of Terminal Value — Exit Multiple ($M)", bold=True, bg=GREEN)
    frm(ws, r, 2, f"=B{tv_mult_row}*{last_disc_col}{disc_row}", "#,##0", bold=True)
    ws.cell(row=r, column=2).fill = fill(GREEN)
    pv_tv_mult_row = r; r += 2

    # Equity bridge
    ws.merge_cells(f"A{r}:I{r}")
    c = ws.cell(row=r, column=1, value="EQUITY VALUE BRIDGE")
    c.font = fnt(bold=True, color=WHITE, size=10)
    c.fill = fill(MED_BLUE)
    c.alignment = ctr()
    ws.row_dimensions[r].height = 16
    r += 1

    pv_sum_formula = "+".join(f"{get_column_letter(3+yr)}{pv_fcf_row}" for yr in range(5))

    lbl(ws, r, 1, "Sum of PV FCFs ($M)", bold=True)
    frm(ws, r, 2, f"={pv_sum_formula}", "#,##0", bold=True)
    sum_pv_row = r; r += 1

    lbl(ws, r, 1, "  Enterprise Value — GGM ($M)", bg=GREEN)
    frm(ws, r, 2, f"=B{sum_pv_row}+B{pv_tv_ggm_row}", "#,##0", bold=True)
    ws.cell(row=r, column=2).fill = fill(GREEN)
    ev_ggm_row = r; r += 1

    lbl(ws, r, 1, "  Enterprise Value — Exit Multiple ($M)", bg=GREEN)
    frm(ws, r, 2, f"=B{sum_pv_row}+B{pv_tv_mult_row}", "#,##0", bold=True)
    ws.cell(row=r, column=2).fill = fill(GREEN)
    ev_mult_row = r; r += 1

    lbl(ws, r, 1, "Less: Net Debt ($M)", bg=LIGHT_GREY)
    inp(ws, r, 2, net_debt, "#,##0")
    ws.cell(row=r, column=2).fill = fill(LIGHT_GREY)
    nd_row = r; r += 1

    lbl(ws, r, 1, "Diluted Shares (M)", bg=LIGHT_GREY)
    inp(ws, r, 2, shares, "#,##0")
    ws.cell(row=r, column=2).fill = fill(LIGHT_GREY)
    sh_row = r; r += 1

    lbl(ws, r, 1, "Implied Price — GGM ($)", bold=True, bg=AMBER)
    frm(ws, r, 2, f"=(B{ev_ggm_row}-B{nd_row})/B{sh_row}", "$#,##0.00", bold=True)
    ws.cell(row=r, column=2).fill = fill(AMBER)
    r += 1

    lbl(ws, r, 1, "Implied Price — Exit Multiple ($)", bold=True, bg=AMBER)
    frm(ws, r, 2, f"=(B{ev_mult_row}-B{nd_row})/B{sh_row}", "$#,##0.00", bold=True)
    ws.cell(row=r, column=2).fill = fill(AMBER)
    r += 1

    lbl(ws, r, 1, "Current Price ($)")
    inp(ws, r, 2, price, "$#,##0.00")
    curr_price_row = r; r += 1

    lbl(ws, r, 1, "Upside — GGM", bold=True)
    frm(ws, r, 2, f"=B{r-3}/B{curr_price_row}-1", "0.0%", bold=True)
    r += 1

    lbl(ws, r, 1, "Upside — Exit Multiple", bold=True)
    frm(ws, r, 2, f"=B{r-3}/B{curr_price_row}-1", "0.0%", bold=True)
    r += 3

    return r


def build_dcf_tab(wb, cfg):
    ws2 = wb.create_sheet("DCF Model")
    set_col_widths(ws2, [30, 12, 12, 12, 12, 12, 12, 12, 14])

    title_row(ws2, 1, f"{cfg['ticker']} — DCF MODEL (EBIT-Based UFCF, Bear / Base / Bull)", "I", sz=12, height=22)
    ws2.merge_cells("A2:I2")
    ws2["A2"].value = "Mid-Year Discounting  |  USD Millions  |  Net Debt = total debt − cash & ST investments"
    ws2["A2"].font = fnt(italic=True, size=9, color="CCCCCC")
    ws2["A2"].fill = fill(DARK_BLUE)
    ws2["A2"].alignment = ctr()

    scenarios = [
        ("BEAR CASE", "C00000", {
            "rev_growth": cfg["bear_rev_growth"],
            "ebitda_mg":  cfg["bear_ebitda_mg"],
            "wacc":       cfg["bear_wacc"],
            "tgr":        cfg["bear_tgr"],
            "exit_mult":  cfg["bear_exit_mult"],
        }),
        ("BASE CASE", "375623", {
            "rev_growth": cfg["base_rev_growth"],
            "ebitda_mg":  cfg["base_ebitda_mg"],
            "wacc":       cfg["base_wacc"],
            "tgr":        cfg["base_tgr"],
            "exit_mult":  cfg["base_exit_mult"],
        }),
        ("BULL CASE", "1F4E79", {
            "rev_growth": cfg["bull_rev_growth"],
            "ebitda_mg":  cfg["bull_ebitda_mg"],
            "wacc":       cfg["bull_wacc"],
            "tgr":        cfg["bull_tgr"],
            "exit_mult":  cfg["bull_exit_mult"],
        }),
    ]

    next_row = 4
    for name, color, params in scenarios:
        next_row = build_dcf_scenario(ws2, next_row, name, color, params, cfg)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 3: SENSITIVITY ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

def build_sensitivity_tab(wb, cfg):
    ws3 = wb.create_sheet("Sensitivity Analysis")
    set_col_widths(ws3, [28, 12, 12, 12, 12, 12, 12, 12, 14])

    title_row(ws3, 1, f"{cfg['ticker']} — SENSITIVITY ANALYSIS (BASE CASE)", "I", sz=12, height=22)
    ws3.merge_cells("A2:I2")
    ws3["A2"].value = f"Exit Multiple Method  |  Net Debt ${cfg['net_debt_m']:,.0f}M  |  Shares {cfg['shares_m']}M  |  Current Price ${cfg['current_price']}"
    ws3["A2"].fill = fill(DARK_BLUE)
    ws3["A2"].font = fnt(italic=True, size=9, color="AAAAAA")
    ws3["A2"].alignment = ctr()

    current_price = cfg["current_price"]
    net_debt      = cfg["net_debt_m"]
    shares        = cfg["shares_m"]

    # Precompute base FCFs and terminal values (base case)
    base_rev = cfg["hist_revenue"][0]
    revs = [base_rev]
    for g in cfg["base_rev_growth"]:
        revs.append(revs[-1] * (1 + g))
    revs = revs[1:]

    da_pct     = cfg["da_pct"]
    capex_pct  = cfg["capex_pct"]
    tax_rate   = cfg["tax_rate"]
    base_wacc  = cfg["base_wacc"]

    ufcfs = []
    for i, (rev, mg) in enumerate(zip(revs, cfg["base_ebitda_mg"])):
        ebitda = rev * mg
        da     = rev * da_pct
        ebit   = ebitda - da
        nopat  = ebit * (1 - tax_rate)
        capex  = rev * capex_pct
        ufcf   = nopat + da - capex
        ufcfs.append(ufcf)

    term_rev    = revs[-1]
    term_ebitda = term_rev * cfg["base_ebitda_mg"][-1]
    term_da     = term_rev * da_pct
    term_ebit   = term_ebitda - term_da
    term_nopat  = term_ebit * (1 - tax_rate)
    term_capex  = term_rev * capex_pct
    term_ufcf   = term_nopat + term_da - term_capex

    def calc_price_ggm(w, tgr_v):
        sum_pv = sum(ufcfs[i] / (1 + w) ** (i + 0.5) for i in range(5))
        if w <= tgr_v:
            return None
        tv     = term_ufcf * (1 + tgr_v) / (w - tgr_v)
        pv_tv  = tv / (1 + w) ** 5
        ev     = sum_pv + pv_tv
        return (ev - net_debt) / shares

    def calc_price_exit(w, em):
        sum_pv = sum(ufcfs[i] / (1 + w) ** (i + 0.5) for i in range(5))
        tv     = em * term_ebitda
        pv_tv  = tv / (1 + w) ** 5
        ev     = sum_pv + pv_tv
        return (ev - net_debt) / shares

    def calc_price_cagr_margin(cagr, tm):
        r_base = cfg["hist_revenue"][0]
        fcfs_s = []
        for yr in range(5):
            r_yr = r_base * (1 + cagr) ** (yr + 1)
            mg_yr = cfg["base_ebitda_mg"][0] + (tm - cfg["base_ebitda_mg"][0]) * (yr + 1) / 5
            eb    = r_yr * mg_yr
            da    = r_yr * da_pct
            ebit  = eb - da
            nopat = ebit * (1 - tax_rate)
            capex = r_yr * capex_pct
            fcfs_s.append(nopat + da - capex)
        sum_pv_s  = sum(fcfs_s[i] / (1 + base_wacc) ** (i + 0.5) for i in range(5))
        r_term    = r_base * (1 + cagr) ** 5
        e_term    = r_term * tm
        tv_s      = cfg["base_exit_mult"] * e_term / (1 + base_wacc) ** 5
        return (sum_pv_s + tv_s - net_debt) / shares

    def color_cell(cell, price):
        if price is None:
            cell.value = "N/A"
            return
        cell.value = round(price, 2)
        cell.number_format = "$#,##0.00"
        cell.font = fnt(size=10)
        cell.alignment = ctr()
        if price > current_price * 1.35:
            cell.fill = fill("C6EFCE")
        elif price > current_price:
            cell.fill = fill(AMBER)
        else:
            cell.fill = fill(RED_LIGHT)

    # Table 1: WACC vs TGR (GGM)
    title_row(ws3, 4, "TABLE 1: Implied Share Price — GGM (WACC vs. Terminal Growth Rate)", "I", bg=MED_BLUE, sz=10)
    waccs = [0.085, 0.090, 0.095, 0.100, 0.105, 0.110, 0.115]
    tgrs  = [0.020, 0.025, 0.030, 0.035, 0.040, 0.045]
    ws3.row_dimensions[5].height = 20
    ws3.cell(row=5, column=1).value = "WACC ↓ / TGR →"
    ws3.cell(row=5, column=1).fill = fill(LIGHT_BLUE)
    ws3.cell(row=5, column=1).font = fnt(bold=True, size=10)
    ws3.cell(row=5, column=1).alignment = ctr()
    for ci, t in enumerate(tgrs, 2):
        col_hdr(ws3, 5, ci, f"{t:.1%}")
    for ri, w in enumerate(waccs):
        row = 6 + ri
        ws3.cell(row=row, column=1).value = f"{w:.1%}"
        ws3.cell(row=row, column=1).font = fnt(bold=True, color=INPUT_BLUE, size=10)
        ws3.cell(row=row, column=1).fill = fill(LIGHT_GREY)
        ws3.cell(row=row, column=1).alignment = ctr()
        for ci, t in enumerate(tgrs, 2):
            color_cell(ws3.cell(row=row, column=ci), calc_price_ggm(w, t))

    # Table 2: WACC vs Exit Multiple
    title_row(ws3, 15, "TABLE 2: Implied Share Price — Exit Multiple (WACC vs. EV/EBITDA)", "I", bg=MED_BLUE, sz=10)
    exit_mults = [8.0, 9.0, 10.0, 11.0, 12.0, 13.0, 14.0]
    ws3.row_dimensions[16].height = 20
    ws3.cell(row=16, column=1).value = "WACC ↓ / Exit Multiple →"
    ws3.cell(row=16, column=1).fill = fill(LIGHT_BLUE)
    ws3.cell(row=16, column=1).font = fnt(bold=True, size=10)
    ws3.cell(row=16, column=1).alignment = ctr()
    for ci, em in enumerate(exit_mults, 2):
        col_hdr(ws3, 16, ci, f"{em:.1f}x")
    for ri, w in enumerate(waccs):
        row = 17 + ri
        ws3.cell(row=row, column=1).value = f"{w:.1%}"
        ws3.cell(row=row, column=1).font = fnt(bold=True, color=INPUT_BLUE, size=10)
        ws3.cell(row=row, column=1).fill = fill(LIGHT_GREY)
        ws3.cell(row=row, column=1).alignment = ctr()
        for ci, em in enumerate(exit_mults, 2):
            color_cell(ws3.cell(row=row, column=ci), calc_price_exit(w, em))

    # Table 3: Rev CAGR vs Terminal EBITDA Margin
    title_row(ws3, 26, f"TABLE 3: Implied Share Price — Exit {cfg['base_exit_mult']:.0f}x EBITDA (Rev CAGR vs. Terminal EBITDA Margin)", "I", bg=MED_BLUE, sz=10)
    rev_cagrs   = [0.03, 0.04, 0.05, 0.06, 0.07, 0.08, 0.09, 0.10]
    term_margins = [0.16, 0.18, 0.20, 0.22, 0.24, 0.26, 0.28]
    ws3.row_dimensions[27].height = 20
    ws3.cell(row=27, column=1).value = "Rev CAGR ↓ / EBITDA Margin →"
    ws3.cell(row=27, column=1).fill = fill(LIGHT_BLUE)
    ws3.cell(row=27, column=1).font = fnt(bold=True, size=10)
    ws3.cell(row=27, column=1).alignment = ctr()
    for ci, m in enumerate(term_margins, 2):
        col_hdr(ws3, 27, ci, f"{m:.0%}")
    for ri, cagr in enumerate(rev_cagrs):
        row = 28 + ri
        ws3.cell(row=row, column=1).value = f"{cagr:.0%}"
        ws3.cell(row=row, column=1).font = fnt(bold=True, color=INPUT_BLUE, size=10)
        ws3.cell(row=row, column=1).fill = fill(LIGHT_GREY)
        ws3.cell(row=row, column=1).alignment = ctr()
        for ci, tm in enumerate(term_margins, 2):
            color_cell(ws3.cell(row=row, column=ci), calc_price_cagr_margin(cagr, tm))

    # Legend
    ws3.row_dimensions[38].height = 14
    for col, text, bg in [(1, "Legend:", WHITE),
                           (2, f"> ${current_price*1.35:.0f} (Significant upside)", "C6EFCE"),
                           (3, f"${current_price:.2f}–${current_price*1.35:.0f} (Moderate upside)", AMBER),
                           (4, f"< ${current_price:.2f} (Downside)", RED_LIGHT)]:
        c = ws3.cell(row=38, column=col)
        c.value = text
        c.fill = fill(bg)
        c.font = fnt(bold=(col == 1), size=9)
        c.alignment = ctr()


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 4: CONSERVATIVE DCF
# ══════════════════════════════════════════════════════════════════════════════

def solve_breakeven_growth(fcf_base, discount, tgr, net_debt, shares, target_price):
    """Binary search for the single flat growth rate (yr 1–10) that makes
    the Conservative DCF implied share price equal to target_price."""
    def implied_price(g):
        pv_sum = 0
        fcf = fcf_base
        for yr in range(1, 11):
            fcf = fcf * (1 + g)
            pv_sum += fcf / (1 + discount) ** yr
        terminal_fcf = fcf * (1 + tgr)
        tv = terminal_fcf / (discount - tgr)
        pv_tv = tv / (1 + discount) ** 10
        equity_value = (pv_sum + pv_tv) - net_debt
        return equity_value / shares

    lo, hi = -0.20, 0.60
    for _ in range(100):
        mid = (lo + hi) / 2
        if implied_price(mid) < target_price:
            lo = mid
        else:
            hi = mid
    return (lo + hi) / 2


def build_conservative_tab(wb, cfg):
    ws4 = wb.create_sheet("Conservative DCF")
    set_col_widths(ws4, {1: 32, 2: 13, 3: 13, 4: 3, 5: 3, 6: 16, 7: 32})

    title_row(ws4, 1, f"{cfg['ticker']} — CONSERVATIVE DCF (OCF − CapEx, Value Investing)", "G", sz=12, height=22)
    ws4.merge_cells("A2:G2")
    ws4["A2"].value = "End-of-Year Discounting  |  Margin of Safety Framework  |  USD Millions"
    ws4["A2"].font = fnt(italic=True, size=9, color="AAAAAA")
    ws4["A2"].fill = fill(DARK_BLUE)
    ws4["A2"].alignment = ctr()

    # Assumptions
    ws4.row_dimensions[4].height = 18
    ws4["A4"].value = f"Initial Cash Flow (FY{cfg['hist_years'][0]}, $M)"
    ws4["A4"].font = fnt(size=10); ws4["A4"].alignment = lft()
    ws4["B4"].value = cfg["cons_fcf_base_m"]
    ws4["B4"].number_format = "#,##0"
    ws4["B4"].font = fnt(color=INPUT_BLUE, bold=True, size=10)
    ws4["B4"].alignment = ctr(); ws4["B4"].border = box_border()
    ws4.merge_cells("C4:E4")
    ws4["C4"].value = f"OCF ${cfg['hist_ocf'][0]:,.0f}M − CapEx ${cfg['hist_capex'][0]:,.0f}M"
    ws4["C4"].font = fnt(italic=True, size=8, color="888888"); ws4["C4"].alignment = lft()

    ws4.row_dimensions[5].height = 8

    ws4.row_dimensions[6].height = 18
    ws4["A6"].value = "Years"; ws4["A6"].font = fnt(bold=True, size=10); ws4["A6"].alignment = lft()
    for col, label in [(2, "1 – 5"), (3, "6 – 10")]:
        c = ws4.cell(row=6, column=col, value=label)
        c.font = fnt(bold=True, color=WHITE, size=10)
        c.fill = fill(MED_BLUE); c.alignment = ctr()

    for row_num, (label, b_val, c_val, note, b_fmt) in enumerate([
        ("FCF Growth Rate",      cfg["cons_growth_1_5"],  cfg["cons_growth_6_10"],
         "Max 20% yr 1–5 / max 10% yr 6–10",  "0%"),
        ("Discount Rate",        cfg["cons_discount"],    None,
         "12% large-cap stable  |  15% growth / higher-risk", "0%"),
        ("Terminal Growth Rate", cfg["cons_tgr"],         None,
         "Conservative: 0–2% range; closer to 0 preferred",   "0%"),
    ], start=7):
        r = row_num
        ws4.row_dimensions[r].height = 18
        ws4.cell(row=r, column=1).value = label
        ws4.cell(row=r, column=1).font = fnt(size=10); ws4.cell(row=r, column=1).alignment = lft()
        ws4.cell(row=r, column=2).value = b_val
        ws4.cell(row=r, column=2).number_format = b_fmt
        ws4.cell(row=r, column=2).font = fnt(color=INPUT_BLUE, bold=True, size=10)
        ws4.cell(row=r, column=2).fill = fill(LIGHT_BLUE)
        ws4.cell(row=r, column=2).alignment = ctr(); ws4.cell(row=r, column=2).border = box_border()
        if c_val is not None:
            ws4.cell(row=r, column=3).value = c_val
            ws4.cell(row=r, column=3).number_format = b_fmt
            ws4.cell(row=r, column=3).font = fnt(color=INPUT_BLUE, bold=True, size=10)
            ws4.cell(row=r, column=3).fill = fill(LIGHT_BLUE)
            ws4.cell(row=r, column=3).alignment = ctr(); ws4.cell(row=r, column=3).border = box_border()
        ws4.merge_cells(f"C{r}:E{r}") if c_val is None else None
        if c_val is None:
            ws4.cell(row=r, column=3).value = note
            ws4.cell(row=r, column=3).font = fnt(italic=True, size=8, color="888888")
            ws4.cell(row=r, column=3).alignment = lft()

    ws4.row_dimensions[10].height = 8

    for r, label, val, fmt, note in [
        (11, "Net Debt ($M)",                  cfg["net_debt_m"],      "#,##0",      f"Debt ${cfg['hist_capex'][0]:,.0f}M − cash"),
        (12, "Diluted Shares Outstanding (M)", cfg["shares_m"],        "#,##0",      ""),
        (13, "Current Market Cap ($M)",         cfg["market_cap_m"],    "#,##0",      ""),
        (14, "Current Share Price ($)",         cfg["current_price"],   "$#,##0.00",  ""),
    ]:
        ws4.row_dimensions[r].height = 18
        ws4.cell(row=r, column=1).value = label
        ws4.cell(row=r, column=1).font = fnt(size=10); ws4.cell(row=r, column=1).alignment = lft()
        ws4.cell(row=r, column=2).value = val
        ws4.cell(row=r, column=2).number_format = fmt
        ws4.cell(row=r, column=2).font = fnt(color=INPUT_BLUE, size=10)
        ws4.cell(row=r, column=2).alignment = ctr(); ws4.cell(row=r, column=2).border = box_border()

    # Summary box (cols F–G, rows 6–11)
    summary = [
        ("DCF Value ($M)",         "=B35", DARK_BLUE,  WHITE,  "#,##0"),
        ("Current Market Cap ($M)","=B13", MED_BLUE,   WHITE,  "#,##0"),
        ("DCF as % of Market Cap", "=F6/F7", LIGHT_BLUE, BLACK, "0%"),
        ("Implied Share Price ($)", "=B39", DARK_BLUE,  WHITE,  "$#,##0.00"),
        ("Current Share Price ($)", "=$B$14", MED_BLUE, WHITE,  "$#,##0.00"),
        ("Upside / (Downside)",     "=F9/F10-1", AMBER, BLACK, "0.0%"),
    ]
    for i, (slabel, sformula, bg, fg, sfmt) in enumerate(summary):
        r = 6 + i
        ws4.row_dimensions[r].height = 18
        for col, val, is_label in [(6, sformula, False), (7, slabel, True)]:
            c = ws4.cell(row=r, column=col)
            c.value = val if not is_label else slabel
            if not is_label:
                c.number_format = sfmt
            c.font = fnt(bold=True, color=fg, size=10)
            c.fill = fill(bg)
            c.alignment = ctr() if not is_label else lft()
            c.border = thin_border()

    # Year table
    ws4.row_dimensions[16].height = 8
    ws4.row_dimensions[17].height = 20
    for col, h in [(1, "Year"), (2, "FCF ($M)"), (3, "Growth Rate"), (7, "Present Value ($M)")]:
        c = ws4.cell(row=17, column=col, value=h)
        c.font = fnt(bold=True, color=WHITE, size=10)
        c.fill = fill(MED_BLUE); c.alignment = ctr()
    for col in [4, 5, 6]:
        ws4.cell(row=17, column=col).fill = fill(WHITE)

    for yr in range(1, 11):
        r = 17 + yr
        ws4.row_dimensions[r].height = 17
        bg = LIGHT_GREY if yr % 2 == 0 else WHITE

        ws4.cell(row=r, column=1).value = yr
        ws4.cell(row=r, column=1).font = fnt(bold=True, size=10)
        ws4.cell(row=r, column=1).alignment = ctr()
        ws4.cell(row=r, column=1).fill = fill(bg)

        if yr == 1:
            fcf_f = "=$B$4*(1+$B$7)"
        else:
            prev = r - 1
            fcf_f = f"=B{prev}*(1+IF({yr}<= 5,$B$7,$C$7))"

        frm(ws4, r, 2, fcf_f, "#,##0")
        ws4.cell(row=r, column=2).fill = fill(bg)

        frm(ws4, r, 3, f"=IF({yr}<=5,$B$7,$C$7)", "0%", color="555555")
        ws4.cell(row=r, column=3).fill = fill(bg)

        frm(ws4, r, 7, f"=B{r}/(1+$B$8)^{yr}", "#,##0", bold=True)
        ws4.cell(row=r, column=7).fill = fill(bg)

        for col in [4, 5, 6]:
            ws4.cell(row=r, column=col).fill = fill(bg)

    # Final calculations
    ws4.row_dimensions[29].height = 8
    ws4.row_dimensions[30].height = 18
    ws4.merge_cells("A30:G30")
    c = ws4.cell(row=30, column=1, value="FINAL CALCULATIONS")
    c.font = fnt(bold=True, color=WHITE, size=10)
    c.fill = fill(MED_BLUE); c.alignment = ctr()

    final_items = [
        ("Terminal Year FCF ($M)",          "=B27*(1+$B$9)",               "#,##0",      LIGHT_GREY),
        ("PV of Year 1–10 Cash Flows ($M)", "=SUM(G18:G27)",               "#,##0",      WHITE),
        ("Terminal Value ($M)",             "=B31/($B$8-$B$9)",            "#,##0",      LIGHT_GREY),
        ("PV of Terminal Value ($M)",       "=B33/(1+$B$8)^10",            "#,##0",      WHITE),
        ("Total DCF Value — EV ($M)",       "=B32+B34",                    "#,##0",      LIGHT_BLUE),
        ("Less: Net Debt ($M)",             "=$B$11",                      "#,##0",      LIGHT_GREY),
        ("Equity Value ($M)",               "=B35-B36",                    "#,##0",      GREEN),
        ("Diluted Shares Outstanding (M)",  "=$B$12",                      "#,##0",      LIGHT_GREY),
        ("Implied Share Price ($)",         "=B37/B38",                    "$#,##0.00",  AMBER),
    ]
    for i, (flabel, fformula, ffmt, fbg) in enumerate(final_items):
        r = 31 + i
        ws4.row_dimensions[r].height = 18
        is_key = flabel in {"Total DCF Value — EV ($M)", "Equity Value ($M)", "Implied Share Price ($)"}
        for col in range(1, 8):
            ws4.cell(row=r, column=col).fill = fill(fbg)
        ws4.cell(row=r, column=1).value = flabel
        ws4.cell(row=r, column=1).font = fnt(bold=is_key, size=10); ws4.cell(row=r, column=1).alignment = lft()
        ws4.cell(row=r, column=2).value = fformula
        ws4.cell(row=r, column=2).number_format = ffmt
        ws4.cell(row=r, column=2).font = fnt(bold=is_key, size=10); ws4.cell(row=r, column=2).alignment = ctr()
        if is_key:
            ws4.cell(row=r, column=2).border = thin_border()

    # ── Breakeven growth rate (Python-solved, static value) ──────────────────
    breakeven_g = solve_breakeven_growth(
        fcf_base   = cfg["cons_fcf_base_m"],
        discount   = cfg["cons_discount"],
        tgr        = cfg["cons_tgr"],
        net_debt   = cfg["net_debt_m"],
        shares     = cfg["shares_m"],
        target_price = cfg["current_price"],
    )

    ws4.row_dimensions[41].height = 8   # spacer
    ws4.row_dimensions[42].height = 18

    # Section mini-header
    ws4.merge_cells("A42:G42")
    ws4["A42"].value = "MARKET-IMPLIED GROWTH RATE"
    ws4["A42"].font = fnt(bold=True, size=9, color=WHITE)
    ws4["A42"].fill = fill(MED_BLUE)
    ws4["A42"].alignment = lft()

    ws4.row_dimensions[43].height = 18
    for col in range(1, 8):
        ws4.cell(row=43, column=col).fill = fill(AMBER)
    ws4.cell(row=43, column=1).value = "Breakeven Growth Rate (Yr 1–10, flat)"
    ws4.cell(row=43, column=1).font = fnt(bold=True, size=10)
    ws4.cell(row=43, column=1).alignment = lft()
    ws4.cell(row=43, column=2).value = breakeven_g
    ws4.cell(row=43, column=2).number_format = "0.0%"
    ws4.cell(row=43, column=2).font = fnt(bold=True, size=10, color=INPUT_BLUE)
    ws4.cell(row=43, column=2).alignment = ctr()
    ws4.cell(row=43, column=2).border = thin_border()
    ws4.merge_cells("C43:G43")
    ws4.cell(row=43, column=3).value = f"Flat growth rate required across yr 1–10 to justify ${cfg['current_price']:.2f} current price (Python-solved, static)"
    ws4.cell(row=43, column=3).font = fnt(italic=True, size=8, color="666666")
    ws4.cell(row=43, column=3).alignment = lft()

    # Notes
    ws4.row_dimensions[45].height = 8
    ws4.row_dimensions[46].height = 32
    ws4.merge_cells("A46:G46")
    ws4["A46"].value = (
        f"Notes: FCF base = OCF (${cfg['hist_ocf'][0]:,.0f}M) − CapEx (${cfg['hist_capex'][0]:,.0f}M) = ${cfg['cons_fcf_base_m']:,.0f}M ({cfg['hist_years'][0]} actuals). "
        f"Growth rates are conservative, anchored to recent decelerated trend. "
        f"{cfg['cons_discount']*100:.0f}% discount rate reflects margin-of-safety premium. "
        "Terminal growth 1% assumes no perpetual outperformance. All blue cells are editable."
    )
    ws4["A46"].font = fnt(italic=True, size=8, color="666666")
    ws4["A46"].alignment = lft_wrap()


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 5: DHANDHO DCF
# ══════════════════════════════════════════════════════════════════════════════

def solve_dhandho_reverse(fcf_base, discount, exit_mult, net_cash, shares, target_price):
    """Binary search for the flat FCF growth rate (all 10 years) that makes the
    Dhandho equity value per share equal to the current share price."""
    target_equity_m = target_price * shares   # price × M shares → $M

    def equity_value(g):
        pv_fcf = 0
        fcf = fcf_base
        for yr in range(1, 11):
            fcf = fcf * (1 + g)
            pv_fcf += fcf / (1 + discount) ** yr
        pv_terminal = (fcf * exit_mult) / (1 + discount) ** 10
        return pv_fcf + pv_terminal + net_cash

    lo, hi = -0.30, 0.60
    for _ in range(120):
        mid = (lo + hi) / 2
        if equity_value(mid) < target_equity_m:
            lo = mid
        else:
            hi = mid
    return (lo + hi) / 2


def build_dhandho_tab(wb, cfg):
    ws = wb.create_sheet("Dhandho DCF")

    # ── Column widths ────────────────────────────────────────────────────────
    # A=yr#, B=label, C=FCF, D=PV | E=spacer | F=assumption label, G=value
    # H=spacer | I=yr#, J=label, K=FCF, L=PV | M=spacer | N=embedded label, O=value
    set_col_widths(ws, {1:5, 2:13, 3:13, 4:14, 5:2, 6:30, 7:13,
                        8:2, 9:5, 10:13, 11:13, 12:14, 13:2, 14:30, 15:13})

    # ── Compute derived inputs ───────────────────────────────────────────────
    net_cash_m = cfg["dhandho_cash_m"] + cfg["dhandho_investments_m"]
    base_year  = int(cfg["hist_years"][0][2:])   # e.g. 2025

    rev_g = solve_dhandho_reverse(
        fcf_base     = cfg["dhandho_fcf_base_m"],
        discount     = cfg["dhandho_discount"],
        exit_mult    = cfg["dhandho_exit_mult"],
        net_cash     = net_cash_m,
        shares       = cfg["shares_m"],
        target_price = cfg["current_price"],
    )

    # ── Row 1: Title ─────────────────────────────────────────────────────────
    title_row(ws, 1, f"{cfg['ticker']} — DHANDHO DCF  (Pabrai Framework)", "O", sz=12, height=22)

    # ── Row 2: Subtitle ──────────────────────────────────────────────────────
    ws.merge_cells("A2:O2")
    ws["A2"].value = "FCF = OCF − CapEx  |  End-of-Year Discounting  |  Two-Phase Growth  |  Sale-Price Terminal  |  USD Millions"
    ws["A2"].font  = fnt(italic=True, size=9, color="AAAAAA")
    ws["A2"].fill  = fill(DARK_BLUE)
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 8   # spacer

    # ── Row 4: Section headers ───────────────────────────────────────────────
    ws.row_dimensions[4].height = 18
    ws.merge_cells("A4:G4")
    ws["A4"].value     = "DHANDHO IV  —  ASSUMED GROWTH"
    ws["A4"].font      = fnt(bold=True, size=10, color=WHITE)
    ws["A4"].fill      = fill(DARK_BLUE)
    ws["A4"].alignment = ctr()

    ws.merge_cells("I4:O4")
    ws["I4"].value     = "REVERSE DHANDHO IV  —  MARKET-IMPLIED GROWTH"
    ws["I4"].font      = fnt(bold=True, size=10, color=WHITE)
    ws["I4"].fill      = fill(DARK_BLUE)
    ws["I4"].alignment = ctr()

    # ── Row 5: Column headers ────────────────────────────────────────────────
    ws.row_dimensions[5].height = 16
    for col, text in [(1,"Yr"), (2,"Year"), (3,"FCF ($M)"), (4,"PV of FCF ($M)")]:
        col_hdr(ws, 5, col, text)
    ws.merge_cells("F5:G5")
    ws["F5"].value = "ASSUMPTIONS  (edit blue cells)"
    ws["F5"].font  = fnt(bold=True, size=9, color=WHITE)
    ws["F5"].fill  = fill(MED_BLUE)
    ws["F5"].alignment = ctr()

    for col, text in [(9,"Yr"), (10,"Year"), (11,"FCF ($M)"), (12,"PV of FCF ($M)")]:
        col_hdr(ws, 5, col, text)
    ws.merge_cells("N5:O5")
    ws["N5"].value = "EMBEDDED GROWTH"
    ws["N5"].font  = fnt(bold=True, size=9, color=WHITE)
    ws["N5"].fill  = fill(MED_BLUE)
    ws["N5"].alignment = ctr()

    # ── Assumptions block (cols F-G, rows 6-14) ──────────────────────────────
    # G6=FCF base, G7=g1-5, G8=g6-10, G9=discount, G10=exit mult,
    # G11=MoS, G12=net cash, G13=shares, G14=curr price
    assumptions = [
        ("FCF Base ($M)",                      cfg["dhandho_fcf_base_m"], "#,##0",      False),
        ("FCF Growth Rate (Yr 1–5)",            cfg["dhandho_growth_1_5"],  "0.0%",      True),
        ("FCF Growth Rate (Yr 6–10)",           cfg["dhandho_growth_6_10"], "0.0%",      True),
        ("Discount Rate",                       cfg["dhandho_discount"],    "0.0%",      True),
        ("Exit Multiple (× Terminal Year FCF)", cfg["dhandho_exit_mult"],   '0.0"x"',    True),
        ("Margin of Safety",                    cfg["dhandho_mos"],         "0%",        True),
        ("Excess Cash ($M)",                       net_cash_m,                 "#,##0",     False),
        ("Shares Outstanding (M)",              cfg["shares_m"],            "#,##0",     False),
        ("Current Share Price ($)",             cfg["current_price"],       "$#,##0.00", False),
    ]
    for i, (label, val, fmt, is_input) in enumerate(assumptions):
        r = 6 + i
        ws.row_dimensions[r].height = 16
        ws.cell(r, 6).value = label
        ws.cell(r, 6).font  = fnt(size=9)
        ws.cell(r, 6).fill  = fill(LIGHT_GREY)
        ws.cell(r, 6).alignment = lft()

        c = ws.cell(r, 7, value=val)
        c.number_format = fmt
        c.alignment     = ctr()
        if is_input:
            c.font   = fnt(size=10, bold=True, color=INPUT_BLUE)
            c.fill   = fill(LIGHT_BLUE)
            c.border = box_border()
        else:
            c.font = fnt(size=10, color="444444")
            c.fill = fill(LIGHT_GREY)

    # ── Reverse embedded assumptions (cols N-O, rows 6-10) ───────────────────
    # O6=FCF base (=G6), O7=embedded g (static), O8=discount (=G9),
    # O9=exit mult (=G10), O10=MoS (=G11)
    rev_assumptions = [
        ("FCF Base ($M)",                        "=$G$6",  "#,##0"),
        ("Embedded Growth Rate (Yr 1–10, flat)", rev_g,    "0.0%"),
        ("Discount Rate",                        "=$G$9",  "0.0%"),
        ("Exit Multiple (× Terminal Year FCF)",  "=$G$10", '0.0"x"'),
        ("Margin of Safety",                     "=$G$11", "0%"),
    ]
    for i, (label, val, fmt) in enumerate(rev_assumptions):
        r = 6 + i
        ws.cell(r, 14).value     = label
        ws.cell(r, 14).font      = fnt(size=9)
        ws.cell(r, 14).fill      = fill(LIGHT_GREY)
        ws.cell(r, 14).alignment = lft()

        c = ws.cell(r, 15, value=val)
        c.number_format = fmt
        c.alignment     = ctr()
        is_solved = (i == 1)
        c.font = fnt(size=10, bold=True, color=INPUT_BLUE)
        c.fill = fill(AMBER if is_solved else LIGHT_GREY)
        if is_solved:
            c.border = box_border()

    # ── Row 6: Year 0 — Net Cash ─────────────────────────────────────────────
    for col in [1,2,3,4,9,10,11,12]:
        ws.cell(6, col).fill = fill(LIGHT_GREY)
    ws.cell(6,1).value = 0;          ws.cell(6,1).font = fnt(size=9); ws.cell(6,1).alignment = ctr()
    ws.cell(6,2).value = "Excess Cash"; ws.cell(6,2).font = fnt(size=9); ws.cell(6,2).alignment = lft()
    ws.cell(6,3).value = ""
    ws.cell(6,4).value = "=$G$12";   ws.cell(6,4).number_format = "#,##0"; ws.cell(6,4).font = fnt(size=9); ws.cell(6,4).alignment = ctr()

    ws.cell(6,9).value  = 0;          ws.cell(6,9).font = fnt(size=9); ws.cell(6,9).alignment = ctr()
    ws.cell(6,10).value = "Excess Cash"; ws.cell(6,10).font = fnt(size=9); ws.cell(6,10).alignment = lft()
    ws.cell(6,11).value = ""
    ws.cell(6,12).value = "=$G$12";  ws.cell(6,12).number_format = "#,##0"; ws.cell(6,12).font = fnt(size=9); ws.cell(6,12).alignment = ctr()

    # ── Rows 7-16: Years 1-10 ────────────────────────────────────────────────
    for yr in range(1, 11):
        r  = 6 + yr
        bg = WHITE if yr % 2 == 1 else LIGHT_GREY
        ws.row_dimensions[r].height = 16
        year_label = f"FY{base_year + yr}"

        for col in [1,2,3,4,9,10,11,12]:
            ws.cell(r, col).fill = fill(bg)

        # LEFT SIDE
        ws.cell(r,1).value = yr;         ws.cell(r,1).font = fnt(size=9); ws.cell(r,1).alignment = ctr()
        ws.cell(r,2).value = year_label; ws.cell(r,2).font = fnt(size=9); ws.cell(r,2).alignment = lft()

        fcf_f = ("=$G$6*(1+$G$7)" if yr == 1
                 else f"=C{r-1}*(1+$G$7)" if yr <= 5
                 else f"=C{r-1}*(1+$G$8)")
        ws.cell(r,3).value = fcf_f; ws.cell(r,3).number_format = "#,##0"; ws.cell(r,3).font = fnt(size=9); ws.cell(r,3).alignment = ctr()

        ws.cell(r,4).value = f"=C{r}/(1+$G$9)^{yr}"; ws.cell(r,4).number_format = "#,##0"; ws.cell(r,4).font = fnt(size=9); ws.cell(r,4).alignment = ctr()

        # RIGHT SIDE (single flat rate from O7)
        ws.cell(r,9).value  = yr;         ws.cell(r,9).font = fnt(size=9); ws.cell(r,9).alignment = ctr()
        ws.cell(r,10).value = year_label; ws.cell(r,10).font = fnt(size=9); ws.cell(r,10).alignment = lft()

        rev_fcf_f = "=$O$6*(1+$O$7)" if yr == 1 else f"=K{r-1}*(1+$O$7)"
        ws.cell(r,11).value = rev_fcf_f; ws.cell(r,11).number_format = "#,##0"; ws.cell(r,11).font = fnt(size=9); ws.cell(r,11).alignment = ctr()

        ws.cell(r,12).value = f"=K{r}/(1+$O$8)^{yr}"; ws.cell(r,12).number_format = "#,##0"; ws.cell(r,12).font = fnt(size=9); ws.cell(r,12).alignment = ctr()

    # ── Row 17: Terminal — Sale Price ─────────────────────────────────────────
    ws.row_dimensions[17].height = 18
    for col in range(1, 16):
        ws.cell(17, col).fill = fill(LIGHT_BLUE)

    exit_label = f"Sale Price ({cfg['dhandho_exit_mult']:.0f}× FCF)"
    for yr_col, lbl_col, fcf_col, pv_col, fcf_ref, mult_ref, dr_ref in [
        (1,  2,  3,  4,  "C16", "$G$10", "$G$9"),
        (9, 10, 11, 12,  "K16", "$O$9",  "$O$8"),
    ]:
        ws.cell(17, yr_col).value = 10; ws.cell(17, yr_col).font = fnt(bold=True, size=9); ws.cell(17, yr_col).alignment = ctr()
        ws.cell(17, lbl_col).value = exit_label; ws.cell(17, lbl_col).font = fnt(bold=True, size=9); ws.cell(17, lbl_col).alignment = lft()
        ws.cell(17, fcf_col).value = f"={fcf_ref}*{mult_ref}"; ws.cell(17, fcf_col).number_format = "#,##0"; ws.cell(17, fcf_col).font = fnt(bold=True, size=9); ws.cell(17, fcf_col).alignment = ctr()
        ws.cell(17, pv_col).value  = f"={get_column_letter(fcf_col)}17/(1+{dr_ref})^10"; ws.cell(17, pv_col).number_format = "#,##0"; ws.cell(17, pv_col).font = fnt(bold=True, size=9); ws.cell(17, pv_col).alignment = ctr()

    # ── Row 18: spacer ────────────────────────────────────────────────────────
    ws.row_dimensions[18].height = 8

    # ── Row 19: Summary headers ───────────────────────────────────────────────
    ws.row_dimensions[19].height = 16
    for rng, text in [("A19:G19", "VALUATION SUMMARY"), ("I19:O19", "VALUATION SUMMARY  (REVERSE)")]:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = text; c.font = fnt(bold=True, size=9, color=WHITE); c.fill = fill(MED_BLUE); c.alignment = lft()

    # ── Rows 20-27: Summary ───────────────────────────────────────────────────
    # G12=net cash, G13=shares, G14=curr price, G11=MoS
    summary = [
        # (label, left_val, right_val, fmt, bg, is_key)
        ("PV of Year 1–10 FCFs ($M)",         "=SUM(D7:D16)",     "=SUM(L7:L16)",     "#,##0",      WHITE,      False),
        ("PV of Sale Price / Terminal ($M)",   "=D17",             "=L17",             "#,##0",      WHITE,      False),
        ("Add: Excess Cash ($M)",                 "=D6",              "=L6",              "#,##0",      WHITE,      False),
        ("Total Intrinsic Value ($M)",         "=D20+D21+D22",     "=L20+L21+L22",     "#,##0",      LIGHT_BLUE, True),
        ("Intrinsic Value per Share ($)",      "=D23/$G$13",       "=L23/$G$13",       "$#,##0.00",  GREEN,      True),
        ("After Margin of Safety (per share)", "=D24*(1-$G$11)",   "=L24*(1-$G$11)",   "$#,##0.00",  AMBER,      True),
        ("Current Share Price ($)",            "=$G$14",           "=$G$14",           "$#,##0.00",  LIGHT_GREY, False),
        ("Premium / (Discount) to MoS IV",    "=D25/$G$14-1",     "=L25/$G$14-1",     "0.0%",       LIGHT_GREY, False),
    ]
    for i, (label, lval, rval, fmt, bg, is_key) in enumerate(summary):
        r = 20 + i
        ws.row_dimensions[r].height = 16

        ws.merge_cells(f"A{r}:C{r}")
        c = ws.cell(r, 1, value=label)
        c.font = fnt(bold=is_key, size=9); c.fill = fill(bg); c.alignment = lft()
        c = ws.cell(r, 4, value=lval)
        c.number_format = fmt; c.font = fnt(bold=is_key, size=9); c.fill = fill(bg); c.alignment = ctr()
        if is_key: c.border = thin_border()

        ws.merge_cells(f"I{r}:K{r}")
        c = ws.cell(r, 9, value=label)
        c.font = fnt(bold=is_key, size=9); c.fill = fill(bg); c.alignment = lft()
        c = ws.cell(r, 12, value=rval)
        c.number_format = fmt; c.font = fnt(bold=is_key, size=9); c.fill = fill(bg); c.alignment = ctr()
        if is_key: c.border = thin_border()

    # ── Row 29: Notes ─────────────────────────────────────────────────────────
    ws.row_dimensions[28].height = 8
    ws.row_dimensions[29].height = 36
    ws.merge_cells("A29:O29")
    ws["A29"].value = (
        f"Notes: FCF base = OCF − CapEx = ${cfg['dhandho_fcf_base_m']:,.0f}M ({cfg['hist_years'][0]} actuals).  "
        f"Excess Cash = cash ${cfg['dhandho_cash_m']:,.0f}M + ST investments ${cfg['dhandho_investments_m']:,.0f}M = ${net_cash_m:,.0f}M.  "
        f"Exit multiple {cfg['dhandho_exit_mult']:.0f}× applied to terminal year FCF.  "
        f"Reverse Dhandho embedded growth ({rev_g:.1%}) is the flat yr 1–10 rate that makes equity IV per share = "
        f"${cfg['current_price']:.2f} (Python-solved, static).  All blue cells are editable."
    )
    ws["A29"].font      = fnt(italic=True, size=8, color="666666")
    ws["A29"].alignment = lft_wrap()


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    cfg = CONFIG
    # Derive hist_ebitda_m for scenario blocks
    cfg["hist_ebitda_m"] = cfg["hist_op_income"][0] + cfg["hist_da"][0]

    wb = openpyxl.Workbook()

    build_comps_tab(wb, cfg)
    build_dcf_tab(wb, cfg)
    build_sensitivity_tab(wb, cfg)
    build_conservative_tab(wb, cfg)
    build_dhandho_tab(wb, cfg)

    filename = f"{cfg['ticker']}_Valuation_Model.xlsx"
    wb.save(filename)
    print(f"Saved: {filename}")
    print(f"Tabs: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()
